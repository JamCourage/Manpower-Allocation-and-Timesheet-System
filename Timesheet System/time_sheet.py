from tkinter import *
from openpyxl import load_workbook, Workbook
from tkcalendar import Calendar
from tkinter.ttk import Combobox
import datetime

import csv
import pandas as pd
import numpy as np

from tkinter import messagebox

def get_work_hours():
    selected_date = cal.get_date()
    selected_level = level_combobox.get()
    selected_module = module_combobox.get()
    selected_project = project_combobox.get()
    hours = hours_entry.get()
    return selected_date, selected_level, selected_module, selected_project, hours

def save_to_excel():
    data = []
    work_hours = get_work_hours()
    # print(work_hours)
    data.append(work_hours)  # 加入data lists

    filename = r"C:\Users\ASUS\Desktop\situation2_scheduling\time_sheet.xlsx"  # 指定固定的檔案名稱

    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Level", "Module", "Project", "Work_hour"])

    for row in data:
        sheet.append(row)

    workbook.save(filename)
    result_label.config(text="資料已儲存至 " + filename)
    
    # 讀取工時表
    time_df = pd.read_excel(r"C:\Users\ASUS\Desktop\situation2_scheduling\time_sheet.xlsx")
    # 將工時表的日期轉成datetime
    for row in range(len(time_df)):
        time_df.loc[row, 'Date'] = datetime.datetime.strptime(time_df.loc[row, 'Date'], "%Y-%m-%d")
    
    # 讀取對應level/month schedule
    input_work_date = datetime.datetime.strptime(work_hours[0], "%Y-%m-%d")
    total_person_day = 0
    
    # 讀對應的scheduling
    file_name = work_hours[1]  # level
    sheet_name = str(input_work_date.month) + '月'  # 幾月
    f = r"C:\Users\ASUS\Desktop\situation2_scheduling\scheduling" + file_name + "_data.xlsx"
    
    scheduling = pd.read_excel(f, sheet_name=sheet_name)
    scheduling = scheduling.set_axis(scheduling.iloc[:, 0], axis=0, copy=False) # 換掉列名稱
    scheduling = scheduling.iloc[:, 1:]  # 移除第一欄
    
    # 建各project還有多少人天沒做
    remain_dict = {}
    
    for pj in scheduling.columns:
        total_person_day = 0
        for row in range(len(time_df)):
            if time_df.loc[row, 'Date'].month == input_work_date.month and \
            time_df.loc[row, 'Level'] == work_hours[1] and \
            time_df.loc[row, 'Module'] == work_hours[2] and \
            time_df.loc[row, 'Project'] == pj:  # 抓到相同月份、level、module、pj
                person_day = (time_df.loc[row, 'Work_hour']) / 8  # 換成人天
                total_person_day += person_day
        remain = scheduling.loc[work_hours[2], pj] - total_person_day
        remain_dict[pj] = remain
    
    # print(remain_dict)
    remain_days = 30 - input_work_date.day
    
    message = f"還剩{remain_days}天，各專案剩餘人天：\n"
    for key, value in remain_dict.items():
        message += f"{key}: {value}\n"
    
    messagebox.showinfo("Notice", message)


###############################################################################
# 創建主視窗
window = Tk()
window.title("工時資訊輸入")
window.geometry("300x400")

# 1日期選擇器
cal = Calendar(window, selectmode="day", date_pattern="yyyy-mm-dd")
cal.pack(pady=10)


# 2Level選擇框
level_label = Label(window, text="Level:")
level_label.pack()
levels = ["L1", "L2", "L3", "L4"]  # 選項列表
level_combobox = Combobox(window, values=levels)
level_combobox.pack()

# 3Module選擇框
module_label = Label(window, text="Module:")
module_label.pack()
modules = ["PM", "PP", "MM", "SD", "CO", "FI"]  # 選項列表
module_combobox = Combobox(window, values=modules)
module_combobox.pack()

# 4Project選擇框
project_label = Label(window, text="Project:")
project_label.pack()
projects = ["專案一", "專案二", "專案三", "專案四", "專案五", "專案六"]  # 選項列表
project_combobox = Combobox(window, values=projects)
project_combobox.pack()

# 5工時輸入欄位
hours_label = Label(window, text="Work_hour:")
hours_label.pack()
hours_entry = Entry(window)
hours_entry.pack()

# 儲存按鈕
save_button = Button(window, text="儲存", command=save_to_excel)
save_button.pack()

# 顯示結果的標籤
result_label = Label(window, text="")
result_label.pack()

# 開始主迴圈
window.mainloop()